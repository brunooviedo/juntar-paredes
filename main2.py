import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl.styles




def cargar_archivo(entry_archivo):
    ruta_archivo = filedialog.askopenfilename(title="Seleccionar archivo")
    entry_archivo.delete(0, tk.END)
    entry_archivo.insert(0, ruta_archivo)

def validar_columnas(df, archivo):
    nombre_columna_2 = df.columns[1] if len(df.columns) > 1 else None

    if nombre_columna_2:
        return True, nombre_columna_2
    else:
        messagebox.showwarning("Columnas faltantes", f"El archivo {archivo} no contiene la segunda columna requerida después de 'Tiempo'.")
        return False, None

def procesar_archivos():
    # Obtener los archivos cargados en los campos de entrada
    archivos = [entry_archivo.get() for entry_archivo in entry_archivos if entry_archivo.get()]

        # Verificar si hay al menos dos archivos cargados
    if len(archivos) < 2:
        messagebox.showwarning("Archivos insuficientes", "Se requieren al menos dos archivos cargados para procesar.")
        return
    
    

     #Crear barra de progreso
    progress_bar = tk.ttk.Progressbar(ventana, orient="horizontal", length=250, mode="determinate")
    progress_bar.grid(row=len(entry_archivos) + 1, column=1, padx=5, pady=5)

    # Crear un nuevo libro de trabajo de Excel
    wb = openpyxl.Workbook()
    initial_sheet = wb["Sheet"]  # Obtener la hoja de cálculo por defecto
    wb.remove(initial_sheet)      # Eliminar la hoja de cálculo por defecto

    # Crear una nueva hoja para los datos sin procesar
    ws_raw = wb.create_sheet(title='Datos sin procesar')


    #itera sobre cada archivo
    for idx, archivo in enumerate(archivos, start=1):
        df = pd.read_excel(archivo)
        #escribir los datos sin procesar en la hoja de cálculo
        for row in df.values:
            row_list = list(row)
            ws_raw.append(row_list)
        # Actualizar barra de progreso
        progress_bar['value'] = (idx / len(archivos)) * 100
        ventana.update_idletasks()
        # Pausa de 0.1 segundos
        ventana.after(100)
    #
    # Crear una nueva hoja para los datos procesados
    ws_processed = wb.create_sheet(title='Datos procesados')

    ultimo_valor = 0

    # Iterar sobre cada archivo
    for idx, archivo in enumerate(archivos, start=1):
        df = pd.read_excel(archivo)
        nombre_columna = df.columns[1]

        df[nombre_columna] += ultimo_valor
        ultimo_valor = df[nombre_columna].iloc[-1]

        for r_idx, row in enumerate(df.values, start=len(ws_processed['B']) + 1):
            for c_idx, value in enumerate(row, start=1):
                ws_processed.cell(row=r_idx, column=c_idx, value=value)
        # Actualizar barra de progreso
        progress_bar['value'] = (idx / len(archivos)) * 100
        ventana.update_idletasks()
        # Pausa de 0.1 segundos
        ventana.after(100)

    for c_idx, header in enumerate(df.columns, start=1):
        ws_processed.cell(row=1, column=c_idx, value=header)

    # Crear un gráfico de dispersión
    chart = ScatterChart()
    chart.title = "Gráfico de Deformación"
    chart.style = 13

    # Tomar la leyenda del valor de la celda B1
    leyenda = ws_processed.cell(row=1, column=2).value

    # Usar los datos a partir de la celda B2
    x_data = Reference(ws_processed, min_col=1, min_row=2, max_row=len(ws_processed['A']))
    y_data = Reference(ws_processed, min_col=2, min_row=2, max_row=len(ws_processed['B']))

    series = Series(y_data, x_data, title=leyenda)
    chart.series.append(series)

    ws_processed.add_chart(chart, "D2")

    ruta_salida = filedialog.asksaveasfilename(title="Guardar archivo Excel", defaultextension=".xlsx", filetypes=[("Archivo de Excel", "*.xlsx")])
    if ruta_salida:
        wb.save(ruta_salida)
        messagebox.showinfo("Proceso completado", "Archivos procesados y guardados correctamente.")

    # Eliminar barra de progreso después de completar el procesamiento
    progress_bar.grid_remove()

def procesar_archivos2():
    try:
        archivos = [entry_archivo.get() for entry_archivo in entry_archivos if entry_archivo.get()]
        if len(archivos) < 1:
            messagebox.showwarning("Archivos insuficientes", "Se requiere al menos un archivo cargado para procesar.")
            return

        # Crear barra de progreso
        progress_bar = tk.ttk.Progressbar(ventana, orient="horizontal", length=250, mode="determinate")
        progress_bar.grid(row=len(entry_archivos) + 1, column=1, padx=5, pady=5)

        # Crear un nuevo libro de trabajo de Excel
        wb = openpyxl.Workbook()
        initial_sheet = wb["Sheet"]  # Obtener la hoja de cálculo por defecto
        wb.remove(initial_sheet)      # Eliminar la hoja de cálculo por defecto

        # Crear una nueva hoja para los datos sin procesar
        ws_raw = wb.create_sheet(title='Datos sin procesar')

        # Lista para almacenar los DataFrames de cada archivo
        df_list = []

        for idx, archivo in enumerate(archivos, start=1):
            try:
                xls = pd.ExcelFile(archivo)
                
                # DataFrame temporal para almacenar los datos del archivo actual
                df_temp = pd.DataFrame()
        
                for sheet_idx, sheet_name in enumerate(xls.sheet_names):
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    
                    # Controlar la inclusión de la columna de tiempo y los encabezados
                    if idx == 1 and sheet_idx == 0:
                        # Añadir la columna de tiempo y la primera columna de datos para la primera hoja del primer archivo
                        df_temp['Tiempo'] = df[df.columns[0]]
                        df_temp[sheet_name] = df[df.columns[1]]
                    else:
                        # Para las hojas siguientes, solo añadir las columnas de datos, omitiendo la de tiempo
                        if 'Tiempo' not in df_temp.columns:
                            df_temp['Tiempo'] = df[df.columns[0]]
                        df_temp = pd.concat([df_temp, pd.DataFrame({sheet_name: df[df.columns[1]]})], axis=1)
        
                # Agregar el DataFrame temporal a la lista de DataFrames
                df_list.append(df_temp)
        
                # Actualizar barra de progreso
                progress_bar['value'] = (idx / len(archivos)) * 100
                ventana.update_idletasks()
                ventana.after(100)
            except ValueError as ve:
                messagebox.showwarning("Archivo sin hojas de cálculo", str(ve))
                continue  # Continuar con el siguiente archivo
            except IndexError:
                messagebox.showwarning("Columna faltante", f"No se encontró la columna en los datos del archivo '{archivo}'.")
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar el archivo '{archivo}': {str(e)}")
                break  # Detiene el procesamiento en caso de error

        # Combinar todos los DataFrames en uno solo
        df_final = pd.concat(df_list, ignore_index=True)
    
        # Guardar los datos en la hoja 'Datos sin procesar'
        for r, row in df_final.iterrows():
            for c, value in enumerate(row, start=1):
                ws_raw.cell(row=r + 1, column=c, value=value)
    
        # Añadir encabezados en la hoja de Excel
        if not df_final.empty:
            for idx, col_name in enumerate(df_final.columns, start=1):
                ws_raw.cell(row=1, column=idx, value=col_name)
    
        # Crear una nueva hoja para los datos procesados
        ws_processed = wb.create_sheet(title='Datos procesados')
    
        # Copiar completamente los primeros valores del primer DataFrame temporal en la hoja 'Datos procesados'
        for r, row in df_list[0].iterrows():
            for c, value in enumerate(row, start=1):
                ws_processed.cell(row=r + 1, column=c, value=value)
    
        # Sumar el último valor del primer DataFrame temporal al primer valor del DataFrame posterior y así sucesivamente
        ultimos_valores = df_list[0].iloc[-1, 1:] if df_list else pd.Series(0, index=df_list[0].columns[1:])
        for df_temp in df_list[1:]:
            for columna in df_temp.columns[1:]:
                df_temp[columna] += ultimos_valores[columna]
                ultimos_valores[columna] = df_temp[columna].iloc[-1]
            # Guardar el DataFrame temporal en la hoja 'Datos procesados' a partir de la segunda fila
            for r, row in enumerate(df_temp.values, start=len(ws_processed['B']) + 1):
                for c, value in enumerate(row, start=1):
                    ws_processed.cell(row=r, column=c, value=value)
    
        # Añadir encabezados en la hoja de Excel para los datos procesados
        for idx, col_name in enumerate(df_final.columns, start=1):
            ws_processed.cell(row=1, column=idx, value=col_name)
    
        # Crear un gráfico de dispersión
        chart = ScatterChart()
        chart.title = "Gráfico de Deformación"
        chart.style = 13
    
        # Usar los datos de la columna A como el eje X
        x_data = Reference(ws_processed, min_col=1, min_row=2, max_row=len(ws_processed['A']))
    
        # Iterar sobre las columnas restantes (excepto la primera) para agregarlas al gráfico
        for col_idx in range(2, len(df_final.columns) + 1):
            # Tomar la leyenda del valor de la celda correspondiente a la columna actual
            leyenda = ws_processed.cell(row=1, column=col_idx).value
    
            # Usar los datos de la columna actual como el eje Y
            y_data = Reference(ws_processed, min_col=col_idx, min_row=2, max_row=len(ws_processed['A']))
    
            # Crear la serie con los datos de la columna actual
            series = Series(y_data, x_data, title=leyenda)
    
            # Agregar la serie al gráfico
            chart.series.append(series)
    
        # Añadir el gráfico a la hoja de cálculo en la celda D2
        ws_processed.add_chart(chart, "G2")
    
            # Guardar el archivo Excel
        ruta_salida = filedialog.asksaveasfilename(title="Guardar archivo Excel", defaultextension=".xlsx", filetypes=[("Archivo de Excel", "*.xlsx")])
        if ruta_salida:
            wb.save(ruta_salida)
            messagebox.showinfo("Proceso completado", "Archivos procesados y guardados correctamente.")
    
        # Eliminar barra de progreso después de completar el procesamiento
        progress_bar.grid_remove()
    except Exception as e:
        messagebox.showerror("Error", f"Error al procesar archivos: {str(e)}, revisa que los archivos contengan las mismas columnas y hojas de calculo")

    
   


def procesar():
    if var_multiples_lineas.get():
        procesar_archivos2()
    else:
        procesar_archivos()

def salir():
    ventana.destroy()

ventana = tk.Tk()
ventana.title("Procesamiento de archivos Excel")

frame_botones = tk.Frame(ventana)
frame_botones.grid(row=11, column=0, columnspan=3, pady=10)

entry_archivos = []
for i in range(10):
    tk.Label(ventana, text=f"Archivo {i+1}:").grid(row=i, column=0, padx=5, pady=5)
    entry_archivo = tk.Entry(ventana, width=50)
    entry_archivo.grid(row=i, column=1, padx=5, pady=5)
    entry_archivos.append(entry_archivo)
    tk.Button(ventana, text="Cargar archivo", command=lambda entry=entry_archivo: cargar_archivo(entry)).grid(row=i, column=2, padx=5, pady=5)

var_multiples_lineas = tk.BooleanVar()
check_multiples_lineas = tk.Checkbutton(ventana, text="Graficar varias Figuras", variable=var_multiples_lineas)
check_multiples_lineas.grid(row=10, column=1, padx=5, pady=5)

btn_procesar = tk.Button(frame_botones, text="Procesar", command=procesar)
btn_salir = tk.Button(frame_botones, text="Salir", command=salir)

btn_procesar.pack(side=tk.LEFT, padx=5)
btn_salir.pack(side=tk.LEFT, padx=5)

frame_botones.grid_configure(sticky="ew")

ventana.mainloop()
