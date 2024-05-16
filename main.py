import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk  # Importar ttk desde tkinter
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
import matplotlib.pyplot as plt

def cargar_archivo(entry_archivo):
    ruta_archivo = filedialog.askopenfilename(title="Seleccionar archivo")
    entry_archivo.delete(0, tk.END)
    entry_archivo.insert(0, ruta_archivo)

def validar_columnas(df, archivo):
    nombre_columna_2 = df.columns[1] if len(df.columns) > 1 else None

    if nombre_columna_2:
        return True, nombre_columna_2
    else:
        messagebox.showwarning("Columnas faltantes", f"El archivo {archivo} no contiene la segunda columna requerida después de 'Tiempo'.")
        return False, None

def procesar_archivos():
    # Obtener los archivos cargados en los campos de entrada
    archivos = [entry_archivo.get() for entry_archivo in entry_archivos if entry_archivo.get()]

        # Verificar si hay al menos dos archivos cargados
    if len(archivos) < 2:
        messagebox.showwarning("Archivos insuficientes", "Se requieren al menos dos archivos cargados para procesar.")
        return

     #Crear barra de progreso
    progress_bar = tk.ttk.Progressbar(ventana, orient="horizontal", length=250, mode="determinate")
    progress_bar.grid(row=len(entry_archivos) + 1, column=1, padx=5, pady=5)

    wb = openpyxl.Workbook()
    initial_sheet = wb["Sheet"]
    wb.remove(initial_sheet)

    ws_raw = wb.create_sheet(title='Datos sin procesar')

    for idx, archivo in enumerate(archivos, start=1):
        df = pd.read_excel(archivo)
        for row in df.values:
            row_list = list(row)
            ws_raw.append(row_list)
        # Actualizar barra de progreso
        progress_bar['value'] = (idx / len(archivos)) * 100
        ventana.update_idletasks()
        # Pausa de 0.1 segundos
        ventana.after(100)

    ws_processed = wb.create_sheet(title='Datos procesados')

    ultimo_valor = 0

    for idx, archivo in enumerate(archivos, start=1):
        df = pd.read_excel(archivo)
        nombre_columna = df.columns[1]

        df[nombre_columna] += ultimo_valor
        ultimo_valor = df[nombre_columna].iloc[-1]

        for r_idx, row in enumerate(df.values, start=len(ws_processed['B']) + 1):
            for c_idx, value in enumerate(row, start=1):
                ws_processed.cell(row=r_idx, column=c_idx, value=value)
        # Actualizar barra de progreso
        progress_bar['value'] = (idx / len(archivos)) * 100
        ventana.update_idletasks()
        # Pausa de 0.1 segundos
        ventana.after(100)

    for c_idx, header in enumerate(df.columns, start=1):
        ws_processed.cell(row=1, column=c_idx, value=header)

    # Crear un gráfico de dispersión
    chart = ScatterChart()
    chart.title = "Gráfico de Deformación"
    chart.style = 13

    # x_data = Reference(ws_processed, min_col=1, min_row=2, max_row=len(ws_processed['A']))
    # y_data = Reference(ws_processed, min_col=2, min_row=2, max_row=len(ws_processed['B']))

    # series = Series(y_data, x_data, title_from_data=True)
    # chart.series.append(series)

    # ws_processed.add_chart(chart, "D2")

        # Tomar la leyenda del valor de la celda B1
    leyenda = ws_processed.cell(row=1, column=2).value

    # Usar los datos a partir de la celda B2
    x_data = Reference(ws_processed, min_col=1, min_row=2, max_row=len(ws_processed['A']))
    y_data = Reference(ws_processed, min_col=2, min_row=2, max_row=len(ws_processed['B']))

    series = Series(y_data, x_data, title=leyenda)
    chart.series.append(series)

    ws_processed.add_chart(chart, "D2")

    ruta_salida = filedialog.asksaveasfilename(title="Guardar archivo Excel", defaultextension=".xlsx", filetypes=[("Archivo de Excel", "*.xlsx")])
    if ruta_salida:
        wb.save(ruta_salida)
        messagebox.showinfo("Proceso completado", "Archivos procesados y guardados correctamente.")

    # Eliminar barra de progreso después de completar el procesamiento
    progress_bar.grid_remove()

def salir():
    ventana.destroy()  # Cerrar la ventana principal y terminar el programa

# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Procesamiento de archivos Excel")

# Crear un marco para contener los botones y centrarlo en la ventana principal
frame_botones = tk.Frame(ventana)
frame_botones.grid(row=11, column=0, columnspan=3, pady=10)  # Asegura que ocupe tres columnas y deja un espacio en la parte inferior

entry_archivos = []
for i in range(10):
    tk.Label(ventana, text=f"Archivo {i+1}:").grid(row=i, column=0, padx=5, pady=5)
    entry_archivo = tk.Entry(ventana, width=50)
    entry_archivo.grid(row=i, column=1, padx=5, pady=5)
    entry_archivos.append(entry_archivo)
    tk.Button(ventana, text="Cargar archivo", command=lambda entry=entry_archivo: cargar_archivo(entry)).grid(row=i, column=2, padx=5, pady=5)

# Crear y colocar los botones dentro del marco
btn_procesar = tk.Button(frame_botones, text="Procesar", command=procesar_archivos)
btn_salir = tk.Button(frame_botones, text="Salir", command=salir)

btn_procesar.pack(side=tk.LEFT, padx=5)  # Colocar el botón "Procesar" a la izquierda con un espacio entre botones
btn_salir.pack(side=tk.LEFT, padx=5)  # Colocar el botón "Salir" a la derecha con un espacio entre botones

# Centrar el marco de botones en la ventana principal
frame_botones.grid_configure(sticky="ew")

ventana.mainloop()
